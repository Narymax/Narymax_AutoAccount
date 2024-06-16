# 常用util
import os
import csv
import tkinter as tk
from tkinter import filedialog
import sys
import xlwt
import os.path as op
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

#  兼容打包后程序的路径读取
def get_current_path():
    if getattr(sys,'frozen',False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    return application_path

# 读取微信、支付宝账单到pandas DataFrame
def read_paylist_file():
    # 打印路径
    current_path = get_current_path()
    print(current_path)
    curpath = current_path
    print(curpath)
    # 读取帐单 文件
    csv_file_path = select_file_from_tk('.csv', show_title='请选择账单文件(支付宝、微信都可以)')
    names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    if csv_file_path == '':
        return 'no_file',None
    else:
        csv_file_path = os.path.abspath(csv_file_path)
        try:
            # 仅用于打开支付宝账单，打开微信、京东账单会报错
            df = pd.read_csv(csv_file_path, names=names, skiprows=0, encoding='gbk')
        except Exception as ex:
            # engine='python' 可以打开，但是中文是乱码 encoding='utf-8' 可以打开，中文正常
            df = pd.read_csv(csv_file_path, names=names, skiprows=0, engine='python', encoding='utf-8')
        return 'read_csv_ok', df


# 通过tkinter选择文件，返回文件路径
# file_extension = ".txt"
def select_file_from_tk(file_extension = '',show_title = '请选择文件'):
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    # file_path = filedialog.askopenfilename()  # 打开文件选择对话框
    curpath = os.path.dirname(os.path.realpath(__file__))
    curpath = get_current_path()

    try:
        if file_extension != '':
            file_path = filedialog.askopenfilename(initialdir=curpath,title=show_title, filetypes=[(file_extension, "*"+file_extension)])
        else:
            file_path = filedialog.askopenfilename(initialdir=curpath,title=show_title)  # 打开文件选择对话框
        # file_path = filedialog.askopenfilename(filetypes=[("csv格式", "*.csv")])  # 打开指定格式
    except Exception as e:
        print("Error selecting file:", e)
        file_path = None

    return file_path


# 屏蔽交易号 （默认不屏蔽手机号）
# 交易号判定，连续数字大于9个,(含有手机号的需要超过11个)
def redacte_key_number(df,colum_name,dedacte_phone_number=False,redacte_trade_number=True,redacte_show_char=''):
        # 含有手机号码的匹配行
        mask = extract_phone(df,colum_name)
        df[mask] = replace_continue_number(df[mask],colum_name,num=12,replace_char=redacte_show_char)

        # 不含有手机号码的行
        df[~mask] = replace_continue_number(df[~mask],colum_name,replace_char=redacte_show_char)
        return df

#  正则提取包含手机的mask
def extract_phone(df,column_name):

        pattern_middle = r'\D1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}\D'  # 匹配中间连续11位数字  aa1731234567bb
        mask_middle = df[column_name].str.contains(pattern_middle, regex=True)
        pattern_begin = r'^1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}\D'  # 匹配从头开始连续11位数字 1731234567aa
        mask_begin = df[column_name].str.contains(pattern_begin, regex=True)
        pattern_end = r'\D1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}$'  # 匹配连续11位数字然后结束  aabb1731234567
        mask_end = df[column_name].str.contains(pattern_end, regex=True)
        mask = mask_middle | mask_begin | mask_end
        print("phone number extracted\n")
        print(df[mask])
        return mask


def replace_continue_number(df, colum_name, num=9, replace_char=''):
    pattern = r'(\d{{{0},}})'.format(num)  # 匹配连续出现num位及以上数字
    # 提取连续出现num位及以上数字的行
    matches = df[colum_name].str.extractall(pattern)
    print(matches)  # 打印匹配的结果

    df[colum_name] = df[colum_name].str.replace(pattern, replace_char)
    print("消除连续出现", num, "位及以上的数字，并且替换成\" ", replace_char, "\"\n")
    return df


def auto_calssify_by_keyword(df,first_classify_col='分类',second_classify_col='子分类', match_list_rule=[], trade_type = '交易类型',redfund_income_classify = True):
    df[first_classify_col] = df[first_classify_col].astype(str)
    df[second_classify_col] = df[second_classify_col].astype(str)
    if match_list_rule is None:
        print('没有找到匹配模板')
        return df
    else:
        outcome_autoclass_total_num = 0
        income_autoclass_total_num = 0
        for list_vob in match_list_rule:

            # 使用列表推导式删除空字符串
            list = [str(item) for item in list_vob if item != '' or item != np.nan]
            first_class = list[1]
            second_class =list[2]

            if len(list) <= 3:
                continue
            else:
                # 支出自动分类
                result = (df['备注'].str.contains(list2orString(list[3:]))) & (df[trade_type].str.contains('支'))
                df.loc[result, first_classify_col] = first_class
                df.loc[result, second_classify_col] = second_class
                modified_outcome_rows = df.loc[result]
                if not modified_outcome_rows.empty:
                    df.loc[result, first_classify_col] = first_class
                    df.loc[result, second_classify_col] = second_class
                    outcome_autoclass_total_num += len(modified_outcome_rows)
                    print("Modified outcome rows:")
                    print(modified_outcome_rows)
        print(f"Total modified outcome rows: {outcome_autoclass_total_num}")
        print("\n \n ")

        # 收入退款分类
        if redfund_income_classify:
            result = df[trade_type].str.contains('收') & df['备注'].str.contains('退款')
            df.loc[result, first_classify_col] = "退款"
            df.loc[result, second_classify_col] = "退款"
            modified_income_rows = df.loc[result]
            if not modified_income_rows.empty:
                df.loc[result, first_classify_col] = "退款"
                df.loc[result, second_classify_col] = "退款"
                income_autoclass_total_num += len(modified_income_rows)
                print("Modified income rows:")
                print(modified_income_rows)

        print(f"Total modified income rows: {income_autoclass_total_num}")
        print("\n \n ")
        return df


def add_count_prefix_character(df, account_name='', account_name2 ='',prefix_character=''):

    if prefix_character != '':
        if account_name != '':
            # df[account_name] = prefix_character + df[account_name]
            # 检查并添加前缀到 account_name 列
            df[account_name] = df[account_name].apply(lambda x: prefix_character + x if x != '' else x)
        if account_name2 != '':
            df[account_name2] = df[account_name2].apply(lambda x: prefix_character + x if x != '' else x)

    return df


# ['He' ,'ui' ,'kk'] -> 'He|ui|kk'
def list2orString(list):
    dstString = ''
    for str in list:
        dstString = dstString + '|' + str
    dstString = dstString[1:]
    return dstString


def check_first_column_contains_string(df, specified_string):
    if df.iloc[:, 0].str.contains(specified_string).any():
        return True
    else:
        return False

def load_html_to_classify_rule_list(info_data,selected_value):
    if selected_value == "随手记":
        print("随手记")

def init_df_columns(df, skiprows=0, use_column_name = True):
    df = df.iloc[skiprows+1:]
    first_row = df.iloc[0]
    # 将第一行转换成字符串
    new_column_names = [str(item) if item != "" else df.columns[idx] for idx, item in enumerate(first_row)]
    # 设置新的列名
    df.columns = new_column_names
    print("导入原始账单")
    print(new_column_names)
    # 删除第一行，因为它已经被用作列名
    df = df.drop(df.index[0])
    # 重置索引
    df.reset_index(drop=True, inplace=True)

    return df

# 使用正则表达式从'账户'列提取"&YYY"格式的字符串
# 将提取出的字符串加到'备注'列的对应行后面
# 从'账户'列删除"&YYY"格式的字符串
def cut_df_Acol_tails_to_Bcol(df,Acol_name='账户',Bcol_name='备注', sep='&'):
    df['extracted'] = df[Acol_name].str.extract(r'(&\w+)$')
    df[Bcol_name] = df[Bcol_name].str.cat(df['extracted'].fillna(''), sep='')
    df[Acol_name] = df[Acol_name].str.replace(r'&\w+$', '')
    df.drop('extracted', axis=1, inplace=True)
    return df

def match_phone_number():
    # 匹配手机号
    # 创建示例数据
    data = {'ID': [1, 2, 3, 4, 5],
            'column_name': ['abc123456789def', '12345678900', 'xyz456789', '123456789012', '15995705685']}
    df = pd.DataFrame(data)

    # 使用正则表达式筛选出含有连续11位数字的行
    pattern = r'^1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}$'  # 匹配连续11位数字
    mask = df['column_name'].str.contains(pattern, regex=True)
    filtered_rows = df[mask]

    # 输出结果
    print(filtered_rows)


def delete_much_than_9_nums():
    # 创建示例数据
    data = {'ID': [1, 2, 3, 4],
            'column_name': ['abc123456789def', '12345678900', 'xyz456789', '123456789012']}
    df = pd.DataFrame(data)

    # 删除连续出现9位及以上数字的行
    pattern = r'\d{9,}'  # 匹配连续出现9位及以上数字
    df['column_name'] = df['column_name'].str.replace(pattern, '')  # 将匹配到的数字替换为空字符串
    filtered_rows = df[df['column_name'].str.len() > 0]  # 筛选出替换后不为空的行

    # 输出结果
    print(filtered_rows)


#  提取包含手机的mask
def extract_phone(df, column_name):
    pattern_middle = r'\D1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}\D'  # 匹配中间连续11位数字  aa1731234567bb
    mask_middle = df[column_name].str.contains(pattern_middle, regex=True)
    pattern_begin = r'^1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}\D'  # 匹配从头开始连续11位数字 1731234567aa
    mask_begin = df[column_name].str.contains(pattern_begin, regex=True)
    pattern_end = r'\D1(3[0-9]|5[0-3,5-9]|7[1-3,5-8]|8[0-9])\d{8}$'  # 匹配连续11位数字然后结束  aabb1731234567
    mask_end = df[column_name].str.contains(pattern_end, regex=True)
    mask = mask_middle | mask_begin | mask_end
    print("phone number extracted\n")
    print(df[mask])
    return mask


def replace_continue_number(df, colum_name, num=9, replace_char=''):
    pattern = r'(\d{{{0},}})'.format(num)  # 匹配连续出现num位及以上数字
    # 提取连续出现num位及以上数字的行
    matches = df[colum_name].str.extractall(pattern)
    print(matches)  # 打印匹配的结果

    df[colum_name] = df[colum_name].str.replace(pattern, replace_char)
    print("消除连续出现", num, "位及以上的数字，并且替换成\" ", replace_char, "\"\n")
    return df


# 屏蔽交易号 （默认不屏蔽手机号）
# 交易号判定，连续数字大于9个,(含有手机号的需要超过11个)
def redacte_key_number(df, colum_name, dedacte_phone_number=False, redacte_trade_number=True, redacte_show_char=''):
    # 含有手机号码的匹配行
    mask = extract_phone(df, colum_name)
    df[mask] = replace_continue_number(df[mask], colum_name, num=12, replace_char=redacte_show_char)

    # 不含有手机号码的行
    df[~mask] = replace_continue_number(df[~mask], colum_name, replace_char=redacte_show_char)
    return df

def convert_back_to_wechat_style_account(df):

    # 确定两列 "商品" "交易对方"
    if "交易信息" in df.columns:
        df["商品"] = df["交易信息"]
    else:
        df["商品"] = ""

    if "交易对方" in df.columns:
        pass
    else:
        df["交易对方"] = ""
    if "当前状态" in df.columns:
        pass
    else:
        df["当前状态"] = ""
    trade_num = ["00000"] * len(df)
    store_num = ["00000"] * len(df)

    # 在备注里面尽量保留更多信息
    df = (
        df.rename(columns={"日期":"交易时间" , "一级分类名称":"交易类型","交易类型":"收/支" , "金额":"金额(元)" ,"账户" :"支付方式" })
        .assign(trade_num=trade_num,store_num=store_num)
        .rename(columns={"trade_num": "交易单号","store_num": "商户单号"})
        .reindex(columns=["交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)", "支付方式", "当前状态", "交易单号", "商户单号", "备注"])
    )

    df["金额(元)"] = '¥' + df["金额(元)"].astype(str)

    df.loc[df['收/支'] == '收入', '当前状态'] = '已存入零钱'
    df.loc[df['收/支'] == '支出', '当前状态'] = '支付成功'
    # 删除交易类型为 “转账” 的数据
    df = df[df['收/支'] != '转账']



    return df

def write_dst_template_file(df, src_name, dst_app_name):
    # 变成指定app 模板
    if dst_app_name == '随手记':
        # 一步完成列名修改、增加新列和重新排序
        df = (
            df.rename(columns={"一级分类名称": "分类", "二级分类名称": "子分类", "账户": "账户1", "*账户": "账户2",
                               "支付渠道": "商家"})
            .reindex(columns=['交易类型', '日期', '分类', '子分类', '账户1', '账户2', '金额', '成员', '商家', '项目',
                              '备注'])
        )
        print("完成 " + dst_app_name + "账单适配")

        file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name +'.xls'

        save_pd_to_xls(df, file_name)
        print("导出文件完成: " + op.join(str(get_current_path()), file_name))

    # 适配钱迹
    if dst_app_name == '钱迹':
        # 定义新列的占位值 账单标记 非必须，可以赋空值
        special_bill_signal_values = [""] * len(df)
        df = (
            df.rename(columns={"日期" : "时间","交易类型":"类型","一级分类名称": "分类", "二级分类名称": "二级分类", "账户": "账户1", "*账户": "账户2"})
            .assign(special_bill_signal=special_bill_signal_values)
            .rename(columns = {"special_bill_signal": "账单标记"})
            .reindex(columns=['时间','分类', '二级分类', '类型', '金额','账户1', '账户2','备注',  '账单标记'])
        )
        # 将日期列从字符串转换为Pandas日期时间对象
        df['时间'] = pd.to_datetime(df['时间'])
        # 将日期时间对象转换为指定格式的字符串（例如：2018-04-08 22:15）
        df['时间'] = df['时间'].dt.strftime('%Y-%m-%d %H:%M')
        # 文件名
        file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '.csv'

        save_pd_to_csv(df, file_name)

        print("保存文件完成: " + op.join(str(get_current_path()), file_name))

    if dst_app_name == '有鱼记账':
        # 定义新列的占位值
        special_billbook = ["日常账本"] * len(df)
        account_type = ["其他资产"]*len(df)
        account_comment = [""] * len(df)
        pic_url = [""] * len(df)
        df = (
            df.rename(columns={"交易类型": "收支类型", "日期": "时间","账户":"资金账户名称","金额":"账目金额", "二级分类名称": "账目分类", "备注":"账目备注"})
            .assign(special_billbook=special_billbook,account_comment=account_comment,account_type=account_type,pic_url=pic_url)
            .rename(columns={"special_billbook": "账本名称","account_type":"资金类型", "account_comment": "资金账户备注","pic_url":"图片"})
            .reindex(columns=['时间', '资金账户名称', '资金类型', '资金账户备注', '收支类型', '账目分类','账目金额','成员','账目备注','账本名称','图片'])
        )
        # 只保留 收支 支出 类型
        df = df[df['收支类型'].isin(['支出', '收入'])]

        file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '.xls'

        save_pd_to_xls(df, file_name, sheet_name="收入支出")
        print("导出文件完成: " + op.join(str(get_current_path()), file_name))

    if dst_app_name == "挖财记账":
        # 定义新列的占位值 添加3列数据
        cruuency_type = ["人民币"] * len(df)
        member_account = [""] * len(df)
        reimbursement = ["非报销"] * len(df)
        payer = [""] * len(df)
        df = (
            df.assign(cruuency_type=cruuency_type,member_account=member_account,reimbursement=reimbursement,payer=payer)
            .rename(columns={"cruuency_type": "币种","member_account":"成员金额","reimbursement":"报销","payer":"付款方"})
        )

        file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '.xls'

        print("导出文件完成: " + op.join(str(get_current_path()), file_name))
        df_outcome = df[df['交易类型'].isin(['支出'])]
        df_outcome = (
            df_outcome.rename(columns={"日期": "消费日期","一级分类名称": "支出大类", "二级分类名称": "支出小类", "金额":"消费金额","项目": "标签", "支付渠道": "商家"})
            .reindex(columns=["消费日期", "支出大类", "支出小类", "消费金额", "币种", "账户", "标签", "商家", "报销", "成员金额", "备注"])

        )
        # save_pd_to_xls(df_outcome, file_name, sheet_name="支出")

        df_income = df[df['交易类型'].isin(['收入'])]
        df_income = (
            df_income.rename(columns={"日期": "收入日期" ,"一级分类名称": "收入大类", "金额":"收入金额","项目": "标签"})
            .reindex(columns=["收入日期", "收入大类", "收入金额", "币种", "账户", "标签", "付款方", "成员金额", "备注", "报销"])
        )
        # save_pd_to_xls(df_income, file_name, sheet_name="收入")

        df_transfer = df[df['交易类型'].isin(['转账'])]
        df_transfer = (
            df_transfer.rename(columns={"日期": "转账时间" ,"账户": "转出账户", "金额":"转出金额","*账户": "转入账户","项目": "标签"})
            .reindex(columns=["转账时间", "转出账户", "转出金额", "币种", "转入账户", "转入金额", "币种", "备注", "标签"])
        )
        # save_pd_to_xls(df_transfer, file_name, sheet_name="转账")

        df_outcome_list,df_income_list,df_transfer_list = df_to_list(df_outcome),df_to_list(df_income),df_to_list(df_transfer)
        # 打开一个工作簿
        workbook = xlwt.Workbook()
        # 添加一个新的工作表
        worksheet = workbook.add_sheet(sheetname="支出", cell_overwrite_ok=True)
        for i, row in enumerate(df_outcome_list):
            for j, value in enumerate(row):
                worksheet.write(i, j, value)

        # 添加一个新的工作表
        worksheet = workbook.add_sheet(sheetname="收入", cell_overwrite_ok=True)
        for i, row in enumerate(df_income_list):
            for j, value in enumerate(row):
                worksheet.write(i, j, value)

        # 添加一个新的工作表
        worksheet = workbook.add_sheet(sheetname="转账", cell_overwrite_ok=True)
        for i,row in enumerate(df_transfer_list):
            for j,value in enumerate(row):
                worksheet.write(i, j, value)
        # 保存工作簿
        workbook.save(op.join(str(get_current_path()), file_name))

    if dst_app_name == "百事AA记账":
        # 定义新列的占位值 添加1列数据 货币
        cruuency_type = ["人民币"] * len(df)
        df = (
            df.assign(cruuency_type=cruuency_type)
            .rename(columns={"cruuency_type": "货币"})
        )
        income_outcome_file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '收支导入模板.xlsx'
        df_inout_come = df[df['交易类型'].isin(['收入','支出'])]
        df_inout_come = (
            df_inout_come.rename(columns={"日期": "账单日期", "交易类型": "账单类型", "二级分类名称": "类别","备注":"描述","支付渠道":"标签"})
            .reindex(columns=["帐单日期", "账单类型", "类别", "金额", "描述", "货币", "账户", "标签"])
        )
        save_pd_to_xlsx(df_inout_come, income_outcome_file_name, sheet_name="payment")
        print("导出文件完成: " + op.join(str(get_current_path()), income_outcome_file_name))




        transfer_file_name = datetime.now().strftime(
            '%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '转账模板.xlsx'
        df_transfer = df[df['交易类型'].isin(['转账'])]
        df_transfer['转出金额']=df_transfer['金额']
        df_transfer['转入金额'] = df_transfer['金额']
        df_transfer['转出币种'] = df_transfer['货币']
        df_transfer['转入币种'] = df_transfer['货币']
        # 定义新列的占位值
        account_book = [""] * len(df_transfer)
        transfer_first_classify = ["转账"] * len(df_transfer)
        transfer_second_classify = [""] * len(df_transfer)
        transfer_info_source = ["自动同步"] * len(df_transfer)
        df_transfer = (
            df_transfer.assign(account_book=account_book,transfer_first_classify=transfer_first_classify,transfer_second_classify=transfer_second_classify,transfer_info_source=transfer_info_source)
            .rename(columns={"日期":"时间","account_book": "账本","账户":"转出账户","*账户":"转入账户", "transfer_first_classify": "大类", "transfer_second_classify": "小类","transfer_info_source":"来源"})
            .reindex(columns=["时间", "账本", "转出账户", "大类", "小类", "转出金额", "转出币种", "转入账户", "转入币种", "转入金额", "备注", "来源"])
        )
        save_pd_to_xlsx(df_transfer,transfer_file_name,sheet_name="转账")
        print("导出文件完成: " + op.join(str(get_current_path()), transfer_file_name))

    # 都不符合 就转回微信格式
    selected_value = dst_app_name
    if (selected_value == "随手记") | (selected_value == "钱迹") | (selected_value == "有鱼记账") | (
                selected_value == "挖财记账") | (selected_value == "百事AA记账"):
        pass
    else:
        df = convert_back_to_wechat_style_account(df)

        # 文件名
        file_name = datetime.now().strftime('%Y-%m-%d %H_%M_%S ') + dst_app_name + '导入' + src_name + '_wechat风格.csv'

        save_pd_to_wx_csv(df, file_name)

        print("保存文件完成: " + op.join(str(get_current_path()), file_name))










    print("")
    return
def save_pd_to_xlsx(df, file_name, sheet_name="sheet1"):
    with pd.ExcelWriter(op.join(str(get_current_path()), file_name), mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def df_to_list(df):
    # 将 DataFrame 转换为二维列表
    return [df.columns.tolist()] + df.values.tolist()

def save_pd_to_csv(df, file_name):
    # 将DataFrame转换为二维的Python列表，并将列名添加到第一行
    df_list = [df.columns.tolist()] + df.astype(str).values.tolist()
    # 创建CSV文件 解决乱码问题
    # 这个编码除了使用UTF-8编码外，还会在文件开头添加一个BOM（Byte Order Mark）标识，这有助于Excel正确地解析文件并避免乱码问题。
    with open(op.join(str(get_current_path()), file_name), mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        for row in df_list:
            writer.writerow(row)

def save_pd_to_wx_csv(df, file_name):
    header_data = [
        ["微信支付账单明细", "", "", "", "", "", "", "", "", "", ""],
        ["微信昵称：[小明]", "", "", "", "", "", "", "", "", "", ""],
        ["起始时间：[2024-04-02 00:00:00] 终止时间：[2024-05-14 14:42:52]", "", "", "", "", "", "", "", "", "", ""],
        ["导出类型：[全部]", "", "", "", "", "", "", "", "", "", ""],
        ["导出时间：[2024-05-14 14:43:30]", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["共999笔记录", "", "", "", "", "", "", "", "", "", ""],
        ["收入：9笔 332.84元", "", "", "", "", "", "", "", "", "", ""],
        ["支出：126笔 9999.99元", "", "", "", "", "", "", "", "", "", ""],
        ["中性交易：0笔 0.00元", "", "", "", "", "", "", "", "", "", ""],
        ["注：", "", "", "", "", "", "", "", "", "", ""],
        ["1. 充值/提现/理财通购买/零钱通存取/信用卡还款等交易，将计入中性交易", "", "", "", "", "", "", "", "", "", ""],
        ["2. 本明细仅展示当前账单中的交易，不包括已删除的记录", "", "", "", "", "", "", "", "", "", ""],
        ["3. 本明细仅供个人对账使用", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["----------------------微信支付账单明细列表--------------------", "", "", "", "", "", "", "", "", "", ""]
    ]
    # 将DataFrame转换为二维的Python列表，并将列名添加到第一行
    df_list = [df.columns.tolist()] + df.astype(str).values.tolist()
    df_dst = header_data + df_list
    # 创建CSV文件 解决乱码问题
    # 这个编码除了使用UTF-8编码外，还会在文件开头添加一个BOM（Byte Order Mark）标识，这有助于Excel正确地解析文件并避免乱码问题。
    with open(op.join(str(get_current_path()), file_name), mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        for row in df_dst:
            writer.writerow(row)

def save_pd_to_xls(df, src_name, sheet_name="Sheet1"):
    with pd.ExcelWriter(op.join(str(get_current_path()), src_name)) as writer:
        # 不保存序号
        df.to_excel(writer, sheet_name=sheet_name, index=False)


# def save_pd_to_xls(df, src_name, sheet_name="Sheet1"):
#     file_path = op.join(str(get_current_path()), src_name)
#
#     if op.exists(file_path):
#         # 如果文件存在，使用 openpyxl 加载工作簿
#         book = load_workbook(file_path)
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
#             writer.book = book
#             # 如果工作表已经存在，删除它
#             if sheet_name in book.sheetnames:
#                 del book[sheet_name]
#             # 写入新的工作表
#             df.to_excel(writer, sheet_name=sheet_name, index=False)
#             writer.save()
#     else:
#         # 如果文件不存在，创建新的文件
#         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#             df.to_excel(writer, sheet_name=sheet_name, index=False)


def print_dog_head():
    print("\n\
                           ::\n\
                          :;J7, :,                        ::;7:\n\
                          ,ivYi, ,                       ;LLLFS:\n\
                          :iv7Yi                       :7ri;j5PL\n\
                         ,:ivYLvr                    ,ivrrirrY2X,\n\
                         :;r@Wwz.7r:                :ivu@kexianli.\n\
                        :iL7::,:::iiirii:ii;::::,,irvF7rvvLujL7ur\n\
                       ri::,:,::i:iiiiiii:i:irrv177JX7rYXqZEkvv17\n\
                    ;i:, , ::::iirrririi:i:::iiir2XXvii;L8OGJr71i\n\
                  :,, ,,:   ,::ir@mingyi.irii:i:::j1jri7ZBOS7ivv,\n\
                     ,::,    ::rv77iiiriii:iii:i::,rvLq@huhao.Li\n\
                 ,,      ,, ,:ir7ir::,:::i;ir:::i:i::rSGGYri712:\n\
               :::  ,v7r:: ::rrv77:, ,, ,:i7rrii:::::, ir7ri7Lri\n\
              ,     2OBBOi,iiir;r::        ,irriiii::,, ,iv7Luur:\n\
            ,,     i78MBBi,:,:::,:,  :7FSL: ,iriii:::i::,,:rLqXv::\n\
            :      iuMMP: :,:::,:ii;2GY7OBB0viiii:i:iii:i:::iJqL;::\n\
           ,     ::::i   ,,,,, ::LuBBu BBBBBErii:i:i:i:i:i:i:r77ii\n\
          ,       :       , ,,:::rruBZ1MBBqi, :,,,:::,::::::iiriri:\n\
         ,               ,,,,::::i:  @arqiao.       ,:,, ,:::ii;i7:\n\
        :,       rjujLYLi   ,,:::::,:::::::::,,   ,:i,:,,,,,::i:iii\n\
        ::      BBBBBBBBB0,    ,,::: , ,:::::: ,      ,,,, ,,:::::::\n\
        i,  ,  ,8BMMBBBBBBi     ,,:,,     ,,, , ,   , , , :,::ii::i::\n\
        :      iZMOMOMBBM2::::::::::,,,,     ,,,,,,:,,,::::i:irr:i:::,\n\
        i   ,,:;u0MBMOG1L:::i::::::  ,,,::,   ,,, ::::::i:i:iirii:i:i:\n\
        :    ,iuUuuXUkFu7i:iii:i:::, :,:,: ::::::::i:i:::::iirr7iiri::\n\
        :     :rk@Yizero.i:::::, ,:ii:::::::i:::::i::,::::iirrriiiri::,\n\
         :      5BMBBBBBBSr:,::rv2kuii:::iii::,:i:,, , ,,:,:i@petermu.,\n\
              , :r50EZ8MBBBBGOBBBZP7::::i::,:::::,: :,:,::i;rrririiii::\n\
                  :jujYY7LS0ujJL7r::,::i::,::::::::::::::iirirrrrrrr:ii:\n\
               ,:  :@kevensun.:,:,,,::::i:i:::::,,::::::iir;ii;7v77;ii;i,\n\
               ,,,     ,,:,::::::i:iiiii:i::::,, ::::iiiir@xingjief.r;7:i,\n\
            , , ,,,:,,::::::::iiiiiiiiii:,:,:::::::::iiir;ri7vL77rrirri::\n\
             :,, , ::::::::i:::i:::i:i::,,,,,:,::i:i:::iir;@Secbone.ii:::")

if __name__ == '__main__':


        data = {'ID': [1, 2, 3, 4,5,6,7],
                'column_name': ['abc123456789def', '12345678900', 'xyz456789', '123456789012','15995705685#154546464875643434','er15995700000sd','01234567']}
        df = pd.DataFrame(data)
        df = redacte_key_number(df,'column_name',redacte_show_char="****")
        print(df)
        print("test")
